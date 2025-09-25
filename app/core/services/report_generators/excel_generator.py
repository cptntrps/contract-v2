"""
Excel Report Generator - Domain Layer

Specialized generator for Excel reports with professional formatting.
Following architectural standards: single responsibility, comprehensive documentation.
"""

from typing import Dict, Any, Optional

from .base_generator import BaseReportGenerator, ReportError
from ....utils.logging.setup import get_logger

# Excel handling with graceful fallback
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

logger = get_logger(__name__)


class ExcelReportGenerator(BaseReportGenerator):
    """
    Specialized Excel report generator with professional formatting.
    
    Purpose: Generates Excel reports with multiple sheets, professional styling,
    and comprehensive change analysis data. Handles Excel-specific features
    like formatting, formulas, and worksheet organization.
    
    AI Context: Excel-specific report generator. Uses openpyxl for Excel file
    creation. When debugging Excel report issues, check dependency availability
    first, then worksheet/formatting logic. Provides multi-sheet reports with
    summary and detailed change data.
    """
    
    @property
    def format_name(self) -> str:
        """Return format name for Excel reports."""
        return 'excel'
    
    @property
    def file_extension(self) -> str:
        """Return file extension for Excel files."""
        return '.xlsx'
    
    @property
    def dependencies_available(self) -> bool:
        """Check if openpyxl is available."""
        return EXCEL_AVAILABLE
    
    def generate_report(self, analysis: Any, output_path: str, 
                       include_styling: bool = True, 
                       include_summary: bool = True,
                       **options) -> bool:
        """
        Generate Excel report with changes table and optional summary sheet.
        
        Purpose: Creates a professional Excel report with multiple worksheets
        containing change analysis data, statistics, and optional AI summary.
        Applies consistent styling and formatting throughout.
        
        Args:
            analysis: Analysis result containing changes and metadata
            output_path: Path to save Excel file
            include_styling: Whether to apply professional styling
            include_summary: Whether to include summary worksheet
            **options: Additional Excel-specific options
            
        Returns:
            True if successful
            
        Raises:
            ReportError: If Excel generation fails or dependencies missing
            
        AI Context: Main Excel report generation method. Creates workbook with
        changes sheet and optional summary sheet. For debugging, check worksheet
        creation, data population, and styling application steps.
        """
        # Validate dependencies and path
        self.validate_dependencies()
        normalized_path = self.validate_output_path(output_path)
        
        self.log_generation_start(normalized_path)
        
        try:
            # Create workbook
            wb = Workbook()
            
            # Generate main changes worksheet
            self._create_changes_worksheet(wb, analysis, include_styling)
            
            # Generate summary worksheet if requested
            if include_summary:
                self._create_summary_worksheet(wb, analysis, include_styling)
            
            # Apply global workbook settings
            self._apply_global_settings(wb, include_styling)
            
            # Save workbook
            wb.save(normalized_path)
            
            self.log_generation_success(normalized_path)
            return True
            
        except Exception as e:
            self.log_generation_failure(normalized_path, e)
            raise ReportError(f"Failed to generate Excel report: {e}")
    
    def _create_changes_worksheet(self, workbook: Any, analysis: Any, 
                                include_styling: bool) -> None:
        """
        Create the main changes worksheet.
        
        Args:
            workbook: Excel workbook object
            analysis: Analysis result data
            include_styling: Whether to apply styling
            
        AI Context: Creates the primary worksheet containing all change data.
        Formats data into table with headers and applies styling if requested.
        """
        ws = workbook.active
        ws.title = "Contract Changes"
        
        # Define headers
        headers = [
            'Change Type', 
            'Original Text', 
            'Modified Text', 
            'Position', 
            'Context',
            'Classification',
            'Explanation'
        ]
        
        # Add headers to first row
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            
            if include_styling:
                self._apply_header_styling(cell)
        
        # Add change data
        changes = getattr(analysis, 'changes', [])
        formatted_changes = self.format_changes_for_display(changes)
        
        for row_idx, change in enumerate(formatted_changes, 2):
            ws.cell(row=row_idx, column=1, value=change['operation'])
            ws.cell(row=row_idx, column=2, value=change['original_text'])
            ws.cell(row=row_idx, column=3, value=change['modified_text'])
            ws.cell(row=row_idx, column=4, value=change['position'])
            ws.cell(row=row_idx, column=5, value=change['context'])
            
            # Add classification and explanation if available
            original_change = changes[row_idx - 2] if row_idx - 2 < len(changes) else None
            if original_change:
                classification = getattr(original_change, 'classification', None)
                if hasattr(classification, 'value'):
                    classification = classification.value
                ws.cell(row=row_idx, column=6, value=classification or '')
                
                ws.cell(row=row_idx, column=7, value=getattr(original_change, 'explanation', '') or '')
            
            # Apply data styling if requested
            if include_styling:
                for col in range(1, len(headers) + 1):
                    self._apply_data_styling(ws.cell(row=row_idx, column=col))
        
        # Adjust column widths
        self._adjust_column_widths(ws, headers)
        
        # Freeze header row if styling is enabled
        if include_styling:
            ws.freeze_panes = 'A2'
    
    def _create_summary_worksheet(self, workbook: Any, analysis: Any, 
                                include_styling: bool) -> None:
        """
        Create summary worksheet with analysis overview.
        
        Args:
            workbook: Excel workbook object
            analysis: Analysis result data
            include_styling: Whether to apply styling
            
        AI Context: Creates secondary worksheet with analysis metadata,
        statistics, and AI summary if available. Provides high-level overview.
        """
        summary_ws = workbook.create_sheet("Analysis Summary")
        
        # Title
        title_cell = summary_ws['A1']
        title_cell.value = "Contract Analysis Summary"
        if include_styling:
            title_cell.font = Font(bold=True, size=16, color="FFFFFF")
            title_cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            title_cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Analysis metadata
        row = 3
        metadata_items = [
            ('Analysis ID', getattr(analysis, 'analysis_id', 'N/A')),
            ('Contract ID', getattr(analysis, 'contract_id', 'N/A')),
            ('Template ID', getattr(analysis, 'template_id', 'N/A')),
            ('Status', getattr(analysis, 'status', 'N/A')),
            ('Created', getattr(analysis, 'created_at', 'N/A'))
        ]
        
        for label, value in metadata_items:
            summary_ws[f'A{row}'] = label
            summary_ws[f'B{row}'] = str(value)
            
            if include_styling:
                summary_ws[f'A{row}'].font = Font(bold=True)
            
            row += 1
        
        # Statistics section
        row += 1
        stats_title_cell = summary_ws[f'A{row}']
        stats_title_cell.value = "Change Statistics"
        if include_styling:
            stats_title_cell.font = Font(bold=True, size=14)
        
        row += 1
        statistics = getattr(analysis, 'statistics', {})
        for key, value in statistics.items():
            summary_ws[f'A{row}'] = key.replace('_', ' ').title()
            summary_ws[f'B{row}'] = value
            row += 1
        
        # LLM analysis section if available
        llm_analysis = getattr(analysis, 'llm_analysis', None)
        if llm_analysis:
            row += 1
            ai_title_cell = summary_ws[f'A{row}']
            ai_title_cell.value = "AI Analysis Summary"
            if include_styling:
                ai_title_cell.font = Font(bold=True, size=14)
            
            row += 1
            summary_text = llm_analysis.get('summary', 'No summary available')
            
            # Wrap long summary text across multiple cells if needed
            wrapped_lines = self.wrap_text(summary_text, 100)
            for line in wrapped_lines:
                summary_ws[f'A{row}'] = line
                if include_styling:
                    summary_ws[f'A{row}'].alignment = Alignment(wrap_text=True, vertical="top")
                row += 1
        
        # Adjust column widths for summary sheet
        summary_ws.column_dimensions['A'].width = 25
        summary_ws.column_dimensions['B'].width = 50
    
    def _apply_header_styling(self, cell: Any) -> None:
        """
        Apply header styling to Excel cell.
        
        Args:
            cell: Excel cell object
            
        AI Context: Consistent header styling used throughout Excel reports.
        Applies professional color scheme and formatting.
        """
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    def _apply_data_styling(self, cell: Any) -> None:
        """
        Apply data cell styling.
        
        Args:
            cell: Excel cell object
        """
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        cell.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    def _adjust_column_widths(self, worksheet: Any, headers: list) -> None:
        """
        Adjust column widths based on content.
        
        Args:
            worksheet: Excel worksheet object
            headers: List of header names
        """
        # Set specific widths for different columns
        width_mapping = {
            'Change Type': 15,
            'Original Text': 30,
            'Modified Text': 30,
            'Position': 12,
            'Context': 40,
            'Classification': 20,
            'Explanation': 35
        }
        
        for col_idx, header in enumerate(headers, 1):
            column_letter = get_column_letter(col_idx)
            width = width_mapping.get(header, 20)
            worksheet.column_dimensions[column_letter].width = width
    
    def _apply_global_settings(self, workbook: Any, include_styling: bool) -> None:
        """
        Apply global workbook settings.
        
        Args:
            workbook: Excel workbook object
            include_styling: Whether to apply styling
        """
        if include_styling:
            # Set default font for all sheets
            for ws in workbook.worksheets:
                ws.sheet_properties.tabColor = "1F77B4"  # Blue tab color
    
    def generate_comparison_report(self, analyses: list, output_path: str, 
                                 **options) -> bool:
        """
        Generate comparison report for multiple analyses.
        
        Args:
            analyses: List of analysis results to compare
            output_path: Path to save comparison report
            **options: Additional options
            
        Returns:
            True if successful
            
        Raises:
            ReportError: If generation fails
            
        AI Context: Specialized method for comparing multiple analysis results
        in a single Excel file. Creates overview table with key metrics.
        """
        self.validate_dependencies()
        normalized_path = self.validate_output_path(output_path)
        
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Analysis Comparison"
            
            # Headers for comparison
            headers = [
                'Analysis ID', 'Contract ID', 'Date', 'Total Changes',
                'Insertions', 'Deletions', 'Replacements', 'Status'
            ]
            
            # Add headers
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col)
                cell.value = header
                self._apply_header_styling(cell)
            
            # Add data for each analysis
            for row, analysis in enumerate(analyses, 2):
                ws.cell(row=row, column=1, value=getattr(analysis, 'analysis_id', ''))
                ws.cell(row=row, column=2, value=getattr(analysis, 'contract_id', ''))
                ws.cell(row=row, column=3, value=str(getattr(analysis, 'created_at', '')))
                
                statistics = getattr(analysis, 'statistics', {})
                ws.cell(row=row, column=4, value=statistics.get('total_changes', 0))
                ws.cell(row=row, column=5, value=statistics.get('insertions', 0))
                ws.cell(row=row, column=6, value=statistics.get('deletions', 0))
                ws.cell(row=row, column=7, value=statistics.get('replacements', 0))
                ws.cell(row=row, column=8, value=getattr(analysis, 'status', ''))
                
                # Apply styling
                for col in range(1, len(headers) + 1):
                    self._apply_data_styling(ws.cell(row=row, column=col))
            
            # Adjust column widths
            self._adjust_column_widths(ws, headers)
            
            # Freeze header row
            ws.freeze_panes = 'A2'
            
            wb.save(normalized_path)
            self.log_generation_success(normalized_path)
            return True
            
        except Exception as e:
            self.log_generation_failure(normalized_path, e)
            raise ReportError(f"Failed to generate Excel comparison report: {e}")